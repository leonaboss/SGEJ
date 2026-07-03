import openpyxl
from django.db import models
from django.http import HttpResponse
from django.utils import timezone


class ImportExportService:
    @staticmethod
    def exportar_modelo_excel(queryset, titulo_hoja, campos):
        """
        campos: lista de tuplas (nombre_campo, etiqueta_encabezado)
        Ej: [('cedula', 'Cédula'), ('nombres', 'Nombres')]
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo_hoja[:31]
        encabezados = [etiqueta for _, etiqueta in campos]
        ws.append(encabezados)
        for obj in queryset:
            fila = []
            for nombre_campo, _ in campos:
                valor = obj
                for parte in nombre_campo.split('__'):
                    if hasattr(valor, parte):
                        valor = getattr(valor, parte)
                    else:
                        valor = ''
                        break
                if callable(valor):
                    valor = valor()
                if valor is None:
                    valor = ''
                elif hasattr(valor, 'strftime'):
                    valor = valor.strftime('%d/%m/%Y %H:%M')
                fila.append(str(valor))
            ws.append(fila)
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"{titulo_hoja.lower().replace(' ', '_')}_{timestamp}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={nombre}'
        wb.save(response)
        return response

    @staticmethod
    def importar_modelo_excel(archivo, modelo_clase, campos, extra_defaults=None):
        """
        campos: lista de tuplas (nombre_campo, etiqueta_encabezado)
        La primera fila del Excel debe contener las etiquetas exactas.
        Los campos virtuales (con __, get_, o callables) se omiten al importar.
        Retorna (creados, errores).
        """
        wb = openpyxl.load_workbook(archivo, read_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        if not filas or len(filas) < 2:
            return 0, ['El archivo está vacío o solo tiene encabezados.']

        encabezados_excel = [str(h).strip() if h else '' for h in filas[0]]
        mapeo = {}
        for nombre_campo, etiqueta in campos:
            if etiqueta in encabezados_excel:
                mapeo[encabezados_excel.index(etiqueta)] = nombre_campo

        # Identificar campos booleanos del modelo
        bool_fields = set()
        for f in modelo_clase._meta.get_fields():
            if isinstance(f, models.BooleanField):
                bool_fields.add(f.name)

        creados = 0
        errores = []
        for idx, fila in enumerate(filas[1:], start=2):
            if not any(fila):
                continue
            try:
                datos = {}
                if extra_defaults:
                    datos.update(extra_defaults)
                for col_idx, nombre_campo in mapeo.items():
                    if '__' in nombre_campo or nombre_campo.startswith('get_'):
                        continue
                    if col_idx < len(fila) and fila[col_idx] is not None:
                        valor = str(fila[col_idx]).strip()
                        if nombre_campo in bool_fields:
                            valor = valor.lower() in ('true', '1', 'sí', 'si', 'yes')
                        datos[nombre_campo] = valor
                if not datos:
                    continue
                modelo_clase.objects.create(**datos)
                creados += 1
            except Exception as e:
                errores.append(f'Fila {idx}: {e}')
        return creados, errores
